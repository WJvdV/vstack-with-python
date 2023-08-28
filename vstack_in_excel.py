'''
Wim van der Veer, augustus 2023
https://github.com/WJvdV
-------------------------------
openpyxl 3.1.0
'''

import openpyxl

wb = openpyxl.load_workbook(filename="vstack_with_python.xlsx", read_only=False)

sheets = wb.sheetnames
rows = wb[sheets[0]].rows
headers = [cell.value for cell in next(rows)]

result = []
i = 0

for sheet in sheets:
    maxCol = wb[ sheets[i] ].max_column
    maxRow = wb[ sheets[i] ].max_row
    for data in wb[sheets[i]].iter_rows(min_row=2, max_col=maxCol, max_row=maxRow, values_only=True):
        result.append(data)
    i = i + 1

wb.create_sheet("combined", 0).append(headers)

for row in result:
    wb['combined'].append(row)

wb.save("vstack_with_python_new.xlsx")
wb.close()
